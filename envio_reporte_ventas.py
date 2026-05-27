import sys
sys.stdout.reconfigure(encoding='utf-8')

import os
from pathlib import Path
from dotenv import load_dotenv
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import io
from datetime import datetime

load_dotenv()

# ─── CONFIGURACIÓN ────────────────────────────────────────────────────────────

BASE_DIR   = Path(__file__).parent
EXCEL_PATH = BASE_DIR / os.getenv("EXCEL_FILE", "Prueba_envio_de_mails.xlsx")
HTML_PATH  = BASE_DIR / os.getenv("HTML_FILE",  "reporte_ventas_vendedor.html")
META       = int(os.getenv("META", 100_000))

SMTP_HOST    = os.getenv("SMTP_HOST", "smtp.gmail.com")
SMTP_PORT    = int(os.getenv("SMTP_PORT", 587))
SMTP_TIMEOUT = int(os.getenv("SMTP_TIMEOUT", 30))  # segundos
SMTP_USER    = os.getenv("SMTP_USER")
SMTP_PASS    = os.getenv("SMTP_PASS")
REMITENTE    = os.getenv("REMITENTE", f"Reportes de Venta <{SMTP_USER}>")

# ─── VALIDACIÓN DE VARIABLES DE ENTORNO ───────────────────────────────────────

_required = {"SMTP_USER": SMTP_USER, "SMTP_PASS": SMTP_PASS}
_missing  = [k for k, v in _required.items() if not v]
if _missing:
    print(f"\nERROR: Variables de entorno requeridas no encontradas: {', '.join(_missing)}")
    print("Crea un archivo .env basándote en .env.example y completa los valores.\n")
    sys.exit(1)

# ─── 1. LEER DATOS ────────────────────────────────────────────────────────────

sheets   = pd.read_excel(EXCEL_PATH, sheet_name=None, dtype={"Vendedor": str})
df_venta = sheets["Ventas"]
df_mails = sheets["Mails"].astype({"Vendedor": str})

# Convertir fecha serial Excel → datetime
df_venta["Fecha"] = pd.to_datetime(df_venta["Fecha"])

# ─── 2. MERGE CON TABLA DE MAILS ──────────────────────────────────────────────

df_venta = df_venta.merge(df_mails, on="Vendedor", how="left")

fecha_max = df_venta["Fecha"].max().strftime("%d/%m/%Y")
fecha_min = df_venta["Fecha"].min().strftime("%d/%m/%Y")

# ─── 3. CARGAR TEMPLATE HTML ──────────────────────────────────────────────────

with open(HTML_PATH, "r", encoding="utf-8") as f:
    html_template = f.read()

# ─── 4. FUNCIÓN: GENERAR EXCEL POR VENDEDOR ───────────────────────────────────

def crear_excel_vendedor(df_det: pd.DataFrame, nombre: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Detalle"

    # Estilos
    header_font    = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill    = PatternFill("solid", start_color="0F172A")
    total_font     = Font(name="Arial", bold=True, size=10)
    total_fill     = PatternFill("solid", start_color="E2E8F0")
    data_font      = Font(name="Arial", size=10)
    center_align   = Alignment(horizontal="center", vertical="center")
    right_align    = Alignment(horizontal="right",  vertical="center")
    left_align     = Alignment(horizontal="left",   vertical="center")
    thin_side      = Side(style="thin", color="CBD5E1")
    thin_border    = Border(bottom=thin_side)

    col_cfg = [
        ("Fecha",    18, "fecha",  left_align),
        ("Folio",    12, "numero", center_align),
        ("Cliente",  30, "texto",  left_align),
        ("Monto",    18, "monto",  right_align),
    ]

    # Titulo
    ws.merge_cells("A1:D1")
    ws["A1"] = f"Detalle de Ventas - {nombre}  |  {fecha_min} al {fecha_max}"
    ws["A1"].font      = Font(name="Arial", bold=True, size=12, color="0F172A")
    ws["A1"].alignment = left_align
    ws.row_dimensions[1].height = 22

    # Encabezados
    for col_idx, (header, width, _, align) in enumerate(col_cfg, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 18

    # Datos
    df_det_sorted = df_det.sort_values("Fecha").reset_index(drop=True)
    for row_idx, row in df_det_sorted.iterrows():
        excel_row = row_idx + 3
        valores = [
            row["Fecha"].strftime("%d/%m/%Y"),
            int(row["Folio"]),
            row["Cliente"],
            row["Monto"],
        ]
        for col_idx, (val, (_, _, tipo, align)) in enumerate(zip(valores, col_cfg), start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=val)
            cell.font      = data_font
            cell.alignment = align
            cell.border    = thin_border
            if tipo == "monto":
                cell.number_format = '#,##0'

    # Fila total
    last_data_row = len(df_det_sorted) + 2
    total_row     = last_data_row + 1
    ws.cell(row=total_row, column=1, value="TOTAL").font      = total_font
    ws.cell(row=total_row, column=1).fill                     = total_fill
    ws.cell(row=total_row, column=1).alignment                = left_align
    ws.merge_cells(f"A{total_row}:C{total_row}")

    total_cell = ws.cell(row=total_row, column=4,
                         value=f"=SUM(D3:D{last_data_row})")
    total_cell.font          = total_font
    total_cell.fill          = total_fill
    total_cell.alignment     = right_align
    total_cell.number_format = '#,##0'

    # Freeze header
    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ─── 5. FUNCIÓN: GENERAR HTML POR VENDEDOR ────────────────────────────────────

def render_html(nombre: str, total: int, meta: int, periodo: str) -> str:
    avance   = total / meta * 100
    avance_s = f"{avance:.1f}%".replace(".", ",")
    color_av = "#10b981" if avance >= 100 else "#f59e0b" if avance >= 75 else "#ef4444"

    html = html_template
    html = html.replace("Vendedor:", f"Vendedor: {nombre}")
    html = html.replace("Abril 2026", periodo)
    html = html.replace("$234.000", f"${total:,.0f}".replace(",", "."))
    html = html.replace("$300.000", f"${meta:,.0f}".replace(",", "."))
    html = html.replace(
        'color:#f59e0b;">78,0%',
        f'color:{color_av};">{avance_s}'
    )
    return html

# ─── 6. FUNCIÓN: GENERAR TEXTO PLANO POR VENDEDOR ─────────────────────────────

def render_plain(nombre: str, total: int, meta: int, periodo: str) -> str:
    """
    Alternativa en texto plano al HTML del email.
    Requerido para evitar que el mensaje vaya a spam (emails solo-HTML
    son penalizados por Gmail y otros clientes).
    """
    avance = total / meta * 100
    lineas = [
        f"Reporte de Ventas | {nombre} | {periodo}",
        "=" * 45,
        "",
        f"Total vendido:  ${total:,.0f}".replace(",", "."),
        f"Meta:           ${meta:,.0f}".replace(",", "."),
        f"Avance:         {avance:.1f}%".replace(".", ","),
        "",
        "Se adjunta el detalle de transacciones en formato Excel.",
        "",
        "--",
        "Reporte generado automaticamente.",
        "Para darse de baja responda con el asunto: BAJA.",
    ]
    return "\n".join(lineas)

# ─── 7. FUNCIÓN: ENVIAR MAIL ──────────────────────────────────────────────────

def enviar_mail(destinatario: str, nombre: str, html_body: str, plain_body: str,
                excel_bytes: bytes, smtp: smtplib.SMTP) -> None:
    """
    Estructura MIME correcta para HTML + texto plano + adjunto:
      mixed
      ├── alternative
      │   ├── text/plain   (fallback)
      │   └── text/html    (preferido)
      └── xlsx (adjunto)
    """
    msg = MIMEMultipart("mixed")
    msg["From"]    = REMITENTE
    msg["To"]      = destinatario
    msg["Subject"] = f"[Reporte Ventas] {nombre} | {fecha_max}"

    # Parte alternativa: texto plano + HTML
    alt = MIMEMultipart("alternative")
    alt.attach(MIMEText(plain_body, "plain", "utf-8"))
    alt.attach(MIMEText(html_body,  "html",  "utf-8"))
    msg.attach(alt)

    # Adjunto Excel
    adjunto = MIMEBase("application",
                       "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    adjunto.set_payload(excel_bytes)
    encoders.encode_base64(adjunto)
    adjunto.add_header("Content-Disposition",
                       f'attachment; filename="Ventas_{nombre}_{fecha_max.replace("/","")}.xlsx"')
    msg.attach(adjunto)

    smtp.sendmail(SMTP_USER, destinatario, msg.as_string())

# ─── 8. LOOP PRINCIPAL ────────────────────────────────────────────────────────

print(f"\n{'─'*55}")
print(f"  Inicio envio: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
print(f"  Periodo: {fecha_min} -> {fecha_max}")
print(f"  Meta individual: ${META:,}")
print(f"{'-'*55}\n")

with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=SMTP_TIMEOUT) as smtp:
    smtp.ehlo()
    smtp.starttls()
    smtp.login(SMTP_USER, SMTP_PASS)

    for _, row_mail in df_mails.iterrows():
        cod    = row_mail["Vendedor"]
        nombre = row_mail["Nombre Vendedor"]
        email  = row_mail["Mails"]

        df_det = df_venta[df_venta["Vendedor"] == cod].copy()
        if df_det.empty:
            print(f"  ⚠  {nombre:<15} - sin datos, se omite")
            continue

        total   = int(df_det["Monto"].sum())
        periodo = df_det["Fecha"].max().strftime("%B %Y").capitalize()

        html_body  = render_html(nombre, total, META, periodo)
        plain_body = render_plain(nombre, total, META, periodo)
        excel_bytes = crear_excel_vendedor(df_det, nombre)

        try:
            enviar_mail(email, nombre, html_body, plain_body, excel_bytes, smtp)
            avance = total / META * 100
            print(f"  ✓  {nombre:<15}  ${total:>12,}  {avance:5.1f}%  ->  {email}")
        except Exception as e:
            print(f"  ✗  {nombre:<15}  ERROR: {e}")

print(f"\n{'-'*55}")
print(f"  Fin: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
print(f"{'-'*55}\n")
